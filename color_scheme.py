from openpyxl.styles import PatternFill

def non_empty(cells):
    return list(filter(lambda cell: cell.value is not None, cells))

def create_cutoff(cutoff, rgb):
    def color(cell):
        if float(cell.value) > cutoff:
            cell.fill = PatternFill('solid', fgColor=rgb)
    def color_all(cells):
        for cell in non_empty(cells):
            color(cell)
    return color_all

def create_match(match, rgb):
    def color(cell):
        if cell.value == match:
            cell.fill = PatternFill('solid', fgColor=rgb)
    def color_all(cells):
        for cell in non_empty(cells):
            color(cell)
    return color_all

def create_tier(thresholds, colors):
    def color(cell, tiers):
        for i, (start, end) in enumerate(tiers):
            value = float(cell.value)
            if ( (start < value and value <= end) or
                 (start == value and value == end)):
                rgb = colors[i]
                cell.fill = PatternFill('solid', fgColor=rgb)
                break

    def color_all(cells):
        assert(len(thresholds) > 0)
        assert(len(thresholds) == len(colors))
        tiers = []
        non_empty_cells = non_empty(cells)
        end = max(map(lambda c: c.value, non_empty_cells))
        threshold = thresholds[0]

        # We are looking for values whose are withing threshold percent diff
        # where percent diff is calculated as
        #   2 * (end-start)/(end+start) < threshold
        # This leads to
        #   (end-start)/(end+start) < threshold/2
        #    end-start              < threshold/2 * (end+start)
        #    end-start              < threshold/2*end + threshold/2*start
        #    end-threshold/2*end    < start+threshold/2*start
        #    end*(1-threshold/2)    < start*(1+threshold/2)
        # and finally
        #    end*(1-threshold/2)/(1+threshold/2)    < start
        start = end * (1-threshold/2) / (1+threshold/2)
        tiers.append((start, end))
        for threshold in thresholds[1:]:
            end = start
            start = end * (1-threshold/2) / (1+threshold/2)
            tiers.append((start, end))

        for cell in non_empty_cells:
            color(cell, tiers)

    return color_all

tier_colors = [ '2EB62C', '83D475', 'C5E8B7' ]
cutoff_color = '2EB62C'
match_color = 'DAAE5D'

default = {
    'Fp (GFLOP/s)': create_tier([0.2, 0.2, 0.2], tier_colors),
    'SB%': create_cutoff(50, cutoff_color),
    'LM%': create_cutoff(50, cutoff_color),
    'RS%': create_cutoff(50, cutoff_color),
    'FE%': create_cutoff(65, cutoff_color),
    'VecType': create_match('SC=100.0%', match_color),
}
